"""
rules_parser.py — Data models and Excel parser for Manual_Testing_Guidelines_Rules.xlsx.

Defines the four sheet-level dataclasses (FeatureRule, EdgeCase, ChecklistItem,
RulesDoc) and the read_rules_excel() function that populates them from the workbook.

Kept separate from rules_to_suites.py so that:
  - The parsing layer has no dependency on generation or step-builder logic.
  - Alternative consumers (reporting, linting) can import models without pulling
    in the full generation pipeline.
"""

from __future__ import annotations

import sys
from dataclasses import dataclass, field

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

from parsers import _str


# --------------------------------------------------------------------------- #
# Data models                                                                  #
# --------------------------------------------------------------------------- #


@dataclass
class FeatureRule:
    feature: str
    logic_item: str
    condition: str
    expected: str


@dataclass
class EdgeCase:
    edge_id: str
    scenario: str
    condition: str
    recovery: str


@dataclass
class ChecklistItem:
    check_id: str
    description: str


@dataclass
class RulesDoc:
    guidelines: list = field(default_factory=list)
    feature_rules: list = field(default_factory=list)
    edge_cases: list = field(default_factory=list)
    checklist: list = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Parser helpers                                                               #
# --------------------------------------------------------------------------- #


def _blank_cell(val) -> bool:
    return val is None or (isinstance(val, str) and not val.strip())


def _header_index(ws) -> dict[str, int]:
    """Map normalised (lowercased, stripped) header text -> column index."""
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if hdr is None:
        return {}
    return {_str(h).lower(): i for i, h in enumerate(hdr) if not _blank_cell(h)}


def _col(row: tuple, idx: int | None, fallback: int) -> str:
    """Read a cell by header index, falling back to a fixed position.

    Header-mapping wins when the column is present (robust to reorder);
    positional fallback keeps working when header text drifts.
    """
    i = idx if idx is not None else fallback
    return _str(row[i]) if 0 <= i < len(row) else ""


# --------------------------------------------------------------------------- #
# Parser                                                                       #
# --------------------------------------------------------------------------- #


def read_rules_excel(path: str) -> RulesDoc:
    """Parse Manual_Testing_Guidelines_Rules.xlsx into a RulesDoc."""
    wb = openpyxl.load_workbook(path, data_only=True)
    doc = RulesDoc()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        hdr = _header_index(ws)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            continue

        if sheet_name == "General Guidelines":
            doc.guidelines = [
                (r[0], r[1], r[2]) for r in rows if any(c is not None for c in r)
            ]

        elif sheet_name == "Feature Rules":
            for r in rows:
                if not any(c is not None for c in r):
                    continue
                doc.feature_rules.append(
                    FeatureRule(
                        feature=_col(r, hdr.get("feature"), 0),
                        logic_item=_col(r, hdr.get("logic item"), 1),
                        condition=_col(r, hdr.get("rule / condition"), 2),
                        expected=_col(r, hdr.get("expected behavior"), 3),
                    )
                )

        elif sheet_name == "Edge Cases":
            for r in rows:
                if not any(c is not None for c in r):
                    continue
                doc.edge_cases.append(
                    EdgeCase(
                        edge_id=_col(r, hdr.get("id"), 0),
                        scenario=_col(r, hdr.get("scenario"), 1),
                        condition=_col(r, hdr.get("condition"), 2),
                        recovery=_col(r, hdr.get("required handling (recovery)"), 3),
                    )
                )

        elif sheet_name == "Release Checklist":
            for r in rows:
                if not any(c is not None for c in r):
                    continue
                doc.checklist.append(
                    ChecklistItem(
                        check_id=_col(r, hdr.get("check id"), 0),
                        description=_col(r, hdr.get("verify item"), 1),
                    )
                )

    return doc
