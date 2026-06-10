"""
excel_tool/parsers.py — Excel sheet parsers and asset validator.

Three parsers (one per sheet) and one validator for IMAGE assets on disk.
All return raw parsed data — no generation logic.
"""

from __future__ import annotations

import math
import os
import re
from typing import Any

import sys

from models import (
    Asset,
    FlowDoc,
    LOCATOR_IMAGE,
    Step,
    ValidationIssue,
)

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")


# ── Helpers ───────────────────────────────────────────────────────────────────


def _blank(val: Any) -> bool:
    if val is None:
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    return isinstance(val, str) and not val.strip()


def _str(val: Any) -> str:
    return "" if _blank(val) else str(val).strip()


def _float(val: Any, default: float) -> float:
    if _blank(val):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _int(val: Any, default: int) -> int:
    if _blank(val):
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


_SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9_.-]+")


def safe_name(name: str) -> str:
    """Sanitise a string for use as a filesystem name."""
    cleaned = _SAFE_NAME_RE.sub("_", name).strip("._")
    return cleaned or "unnamed"


# ── Header Helpers ────────────────────────────────────────────────────────────


def _read_headers(ws):
    row = next(ws.iter_rows(values_only=True), None)
    if row is None:
        return [], "Sheet is empty"
    return [_str(h).lower() for h in row], None


def _row_map(headers: list[str], row: tuple) -> dict[str, Any]:
    return {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}


# ── Sheet Name Constants ──────────────────────────────────────────────────────

DEFAULT_OBJECTS_SHEET = "Object_Repository"
DEFAULT_ACTIONS_SHEET = "Action_Logic"
DEFAULT_STEPS_SHEET = "Test_Execution"

_REQUIRED_OBJ_HEADERS = {"object_id"}
_REQUIRED_STEP_HEADERS = {"suite_id", "step", "action_keyword"}
_REQUIRED_ACTION_HEADERS = {"logic_id", "action_name"}


# ── Object Repository Parser ──────────────────────────────────────────────────


def parse_object_repository(
    wb, sheet_name: str = DEFAULT_OBJECTS_SHEET
) -> tuple[dict[str, Asset], list[str]]:
    """Parse the Object_Repository sheet into ``{object_id: Asset}``."""
    if sheet_name not in wb.sheetnames:
        return {}, [f"Sheet '{sheet_name}' not found"]

    ws = wb[sheet_name]
    headers, err = _read_headers(ws)
    if err:
        return {}, [f"{sheet_name}: {err}"]

    missing = _REQUIRED_OBJ_HEADERS - set(headers)
    if missing:
        return {}, [f"{sheet_name} missing required columns: {missing}"]

    assets: dict[str, Asset] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(_blank(c) for c in row):
            continue
        r = _row_map(headers, row)
        name = _str(r.get("object_id"))
        if not name:
            continue
        assets[name] = Asset(
            object_id=name,
            locator_type=_str(r.get("locator_type")).upper() or LOCATOR_IMAGE,
            resource_path=_str(r.get("resource_path")),
            threshold=_float(r.get("smart_threshold"), 0.8),
            timeout=_float(r.get("timeout"), 5.0),
        )
    return assets, []


# ── Action Logic Parser ───────────────────────────────────────────────────────


def parse_action_logic(
    wb, sheet_name: str = DEFAULT_ACTIONS_SHEET
) -> tuple[set[str], list[FlowDoc], list[str]]:
    """Parse the Action_Logic sheet.

    Returns:
        (known_keywords, flow_descriptions, errors)
    """
    if sheet_name not in wb.sheetnames:
        return set(), [], []

    ws = wb[sheet_name]
    headers, err = _read_headers(ws)
    if err:
        return set(), [], [f"{sheet_name}: {err}"]

    missing = _REQUIRED_ACTION_HEADERS - set(headers)
    if missing:
        return set(), [], [f"{sheet_name} missing required columns: {missing}"]

    keywords: set[str] = set()
    flows: list[FlowDoc] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(_blank(c) for c in row):
            continue
        r = _row_map(headers, row)
        logic_id = _str(r.get("logic_id"))
        name = _str(r.get("action_name"))
        if not logic_id or not name:
            continue
        if logic_id.upper().startswith("ACT"):
            keywords.add(name.upper())
        else:
            flows.append(
                FlowDoc(
                    logic_id=logic_id,
                    action_name=name,
                    command=_str(r.get("machine_command")),
                    target_page=_str(r.get("target_page")),
                )
            )
    return keywords, flows, []


# ── Test Execution Parser ─────────────────────────────────────────────────────


def parse_test_execution(
    wb, sheet_name: str = DEFAULT_STEPS_SHEET
) -> tuple[dict[str, list[Step]], list[str]]:
    """Parse the Test_Execution sheet into ``{suite_id: [Step, ...]}``."""
    if sheet_name not in wb.sheetnames:
        return {}, [f"Sheet '{sheet_name}' not found"]

    ws = wb[sheet_name]
    headers, err = _read_headers(ws)
    if err:
        return {}, [f"{sheet_name}: {err}"]

    missing = _REQUIRED_STEP_HEADERS - set(headers)
    if missing:
        return {}, [f"{sheet_name} missing required columns: {missing}"]

    suites: dict[str, list[Step]] = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(_blank(c) for c in row):
            continue
        r = _row_map(headers, row)
        suite = _str(r.get("suite_id"))
        if not suite:
            continue
        suites.setdefault(suite, []).append(
            Step(
                suite_id=suite,
                step_no=_int(r.get("step"), 0),
                action=_str(r.get("action_keyword")).upper(),
                excel_row=row_num,
                target=_str(r.get("target_id")),
                params=_str(r.get("params")),
                expected=_str(r.get("expected_result")),
            )
        )
    return suites, []


# ── Asset Validator ───────────────────────────────────────────────────────────


def validate_assets(assets: dict[str, Asset], base_dir: str) -> list[ValidationIssue]:
    """Check that every IMAGE asset's resource path exists on disk."""
    issues: list[ValidationIssue] = []
    for name, asset in assets.items():
        if asset.locator_type != LOCATOR_IMAGE:
            continue
        path = asset.resource_path
        if not path or path.upper() == "NONE":
            continue
        full = path if os.path.isabs(path) else os.path.join(base_dir, path)
        if not os.path.isfile(full):
            issues.append(ValidationIssue(component=name, path=path))
            print(f"  [WARN] Asset not found on disk: {path}  (object: {name})")
    return issues
