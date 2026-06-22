#!/usr/bin/env python3
"""Step 1: Load and Verify Excel workbook."""

import sys
import json
import openpyxl

REQUIRED_SHEETS = ["Object_Repository", "Action_Logic", "Test_Execution"]


def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else None
    if not excel_path:
        print("Usage: python agent_step1.py <path-to-workbook.xlsx>")
        sys.exit(1)
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Could not open workbook: {e}")
        sys.exit(1)

    print(f"File loaded: {excel_path}")
    print(f"Sheets found: {wb.sheetnames}")

    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        print(f"ERROR: Missing mandatory sheets: {missing}")
        sys.exit(1)

    stats = {}
    for sheet_name in REQUIRED_SHEETS:
        ws = wb[sheet_name]
        # Count data rows (rows after header row 1 that have at least one non-None cell)
        data_rows = 0
        for row in ws.iter_rows(min_row=2):
            if any(cell.value is not None for cell in row):
                data_rows += 1
        stats[sheet_name] = {
            "max_row": ws.max_row,
            "data_rows": data_rows,
            "max_col": ws.max_column,
        }
        print(
            f"  Sheet '{sheet_name}': max_row={ws.max_row}, data_rows={data_rows}, max_col={ws.max_column}"
        )

    # Print header row for each sheet
    for sheet_name in REQUIRED_SHEETS:
        ws = wb[sheet_name]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"  Headers '{sheet_name}': {headers}")

    print("VERIFICATION: All mandatory sheets present.")
    print(json.dumps(stats))


if __name__ == "__main__":
    main()
