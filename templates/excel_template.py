"""
templates/excel_template.py — generate template.xlsx for testers.

Writes Object_Repository, Action_Logic, and Test_Execution sheets with
pre-filled headers, example rows, and an Action_Keyword dropdown validation.
"""

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")


ALL_KEYWORDS = [
    "TAP",
    "TOUCH",
    "WAIT_FOR",
    "ASSERT_VISIBLE",
    "START_APP",
    "STOP_APP",
    "INPUT_TEXT",
    "READ_TEXT",
    "SWIPE",
    "SCROLL",
    "LONG_PRESS",
    "SLEEP",
    "BACK",
    "HOME",
    "SNAPSHOT",
]


def generate_template(output_path: str = "template.xlsx") -> str:
    """Write a filled template.xlsx and return the absolute output path."""
    wb = openpyxl.Workbook()
    ws_obj = wb.active
    ws_obj.title = "Object_Repository"
    _write_object_repository(ws_obj)

    ws_act = wb.create_sheet("Action_Logic")
    _write_action_logic(ws_act)

    ws_test = wb.create_sheet("Test_Execution")
    _write_test_execution(ws_test)

    wb.save(output_path)
    wb.close()
    return os.path.abspath(output_path)


def _header_row(ws, headers):
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _write_object_repository(ws):
    headers = [
        "Object_ID",
        "Locator_Type",
        "Resource_Path",
        "Smart_Threshold",
        "Timeout",
    ]
    _header_row(ws, headers)
    examples = [
        ("btn_play", "IMAGE", "./assets/home/btn_play.png", 0.85, 5),
        ("heart_count", "OCR", "NONE", 0.70, 10),
        ("btn_close", "IMAGE", "./assets/popup/btn_close.png", 0.90, 3),
    ]
    for i, row in enumerate(examples, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)
    for col, width in zip("ABCDE", [20, 12, 42, 16, 8]):
        ws.column_dimensions[col].width = width


def _write_action_logic(ws):
    headers = ["Logic_ID", "Action_Name", "Machine_Command", "Target_Page"]
    _header_row(ws, headers)
    for i, kw in enumerate(ALL_KEYWORDS, 1):
        row = (f"ACT_{i:03d}", kw, f"{kw}(Object_ID)", "")
        for j, val in enumerate(row, 1):
            ws.cell(row=i + 1, column=j, value=val)
    for col, width in zip("ABCD", [10, 16, 30, 20]):
        ws.column_dimensions[col].width = width


def _write_test_execution(ws):
    headers = [
        "Suite_ID",
        "Step",
        "Action_Keyword",
        "Target_ID",
        "Params",
        "Expected_Result",
    ]
    _header_row(ws, headers)
    examples = [
        (
            "TC_EXAMPLE_LOGIN",
            1,
            "START_APP",
            "",
            '{"heart":"5","level":"1"}',
            "App opens",
        ),
        ("TC_EXAMPLE_LOGIN", 2, "WAIT_FOR", "btn_play", "", "Home screen loaded"),
        ("TC_EXAMPLE_LOGIN", 3, "TAP", "btn_play", "", "Game starts"),
        ("TC_EXAMPLE_LOGIN", 4, "SLEEP", "", "2", "Wait for animation"),
        ("TC_EXAMPLE_LOGIN", 5, "SNAPSHOT", "", "after_tap", "Screenshot saved"),
    ]
    for i, row in enumerate(examples, 2):
        for j, val in enumerate(row, 1):
            ws.cell(row=i, column=j, value=val)

    formula = '"' + ",".join(ALL_KEYWORDS) + '"'
    dv = DataValidation(
        type="list", formula1=formula, allow_blank=False, showDropDown=False
    )
    dv.error = "Must be a valid action keyword"
    dv.errorTitle = "Invalid Keyword"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add("C2:C1000")

    for col, width in zip("ABCDEF", [24, 6, 16, 20, 38, 25]):
        ws.column_dimensions[col].width = width
