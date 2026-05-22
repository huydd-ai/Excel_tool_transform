"""
excel_to_airtest.py — AutomationRebase Excel -> Airtest .air script generator

Sheets read from the workbook (canonical schema, case-insensitive headers):

  Object_Repository
    Object_ID | Locator_Type | Resource_Path | Smart_Threshold | Timeout
      (Locator_Type: IMAGE | OCR; OCR rows currently emit TODO stubs)

  Action_Logic            (optional - used for keyword registry + flow reference)
    Logic_ID | Action_Name | Machine_Command | Target_Page
      (ACT_* rows declare primitives; FLOW_* rows are composite descriptions)

  Test_Execution
    Suite_ID | Step | Action_Keyword | Target_ID | Params | Expected_Result

One Suite_ID -> one .air script. The same workbook can describe many test cases
across many projects without code changes; per-project differences (e.g. app
package id, framework imports, composite-flow expansion, page object routing)
come from a single subclass of AirtestGenerator.

Architecture:
  Parser    -> dataclasses  (generator-based, low memory)
  Validator -> disk check for IMAGE assets
  Generator -> AirtestGenerator class — subclass and override hooks/handlers
  Writer    -> .air dir per suite + structured report

Per-project extension:
  Drop a one-file subclass into generators/<your_project>.py:

      from excel_to_airtest import AirtestGenerator, action

      class MyGenerator(AirtestGenerator):
          IMPORTS = "from my_framework import *"
          DEFAULT_APP_PACKAGE = "com.my.app"

          @action("CLICK")
          def handle_click(self, step, ctx):
              asset, lines, issue = self._resolve_image(step, ctx)
              if asset is None: return lines, issue
              return [f"my_framework.tap({asset.resource_path!r})"], None

      if __name__ == "__main__":
          MyGenerator.main()

Security: all Excel string values embedded in generated code pass through
repr() / !r - prevents code injection via crafted cell content.
"""

import argparse
import importlib
import json
import math
import os
import re
import sys
from dataclasses import dataclass
from typing import Callable

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")


# --------------------------------------------------------------------------- #
# Data model                                                                  #
# --------------------------------------------------------------------------- #

LOCATOR_IMAGE = "IMAGE"
LOCATOR_OCR   = "OCR"

_SCROLL_PRESETS = {
    "up":    ("(int(w*0.5), int(h*0.7))", "(int(w*0.5), int(h*0.3))"),
    "down":  ("(int(w*0.5), int(h*0.3))", "(int(w*0.5), int(h*0.7))"),
    "left":  ("(int(w*0.7), int(h*0.5))", "(int(w*0.3), int(h*0.5))"),
    "right": ("(int(w*0.3), int(h*0.5))", "(int(w*0.7), int(h*0.5))"),
}


@dataclass
class Asset:
    object_id:     str
    locator_type:  str   = LOCATOR_IMAGE
    resource_path: str   = ""
    threshold:     float = 0.8
    timeout:       float = 5.0


@dataclass
class Step:
    suite_id:  str
    step_no:   int
    action:    str
    excel_row: int = 0
    target:    str = ""
    params:    str = ""
    expected:  str = ""


@dataclass
class FlowDoc:
    logic_id:    str
    action_name: str
    command:     str
    target_page: str


@dataclass
class ValidationIssue:
    component: str
    path:      str
    kind:      str = "FILE_NOT_FOUND"


@dataclass
class GenerationIssue:
    suite_id:  str
    step_no:   int
    excel_row: int
    reason:    str


@dataclass
class GenCtx:
    assets:      dict
    app_package: str


# --------------------------------------------------------------------------- #
# Helpers                                                                     #
# --------------------------------------------------------------------------- #

def _blank(val) -> bool:
    if val is None:
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    return isinstance(val, str) and not val.strip()


def _str(val) -> str:
    return "" if _blank(val) else str(val).strip()


def _float(val, default: float) -> float:
    if _blank(val):
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _int(val, default: int) -> int:
    if _blank(val):
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


_SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9_.-]+")


def _safe_name(name: str) -> str:
    cleaned = _SAFE_NAME_RE.sub("_", name).strip("._")
    return cleaned or "unnamed"


# --------------------------------------------------------------------------- #
# Parser                                                                      #
# --------------------------------------------------------------------------- #

_REQUIRED_OBJ_HEADERS    = {"object_id"}
_REQUIRED_STEP_HEADERS   = {"suite_id", "step", "action_keyword"}
_REQUIRED_ACTION_HEADERS = {"logic_id", "action_name"}


def _read_headers(ws):
    row = next(ws.iter_rows(values_only=True), None)
    if row is None:
        return [], "Sheet is empty"
    return [_str(h).lower() for h in row], None


def _row_map(headers, row):
    return {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}


def parse_object_repository(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return {}, [f"Sheet '{sheet_name}' not found"]

    ws = wb[sheet_name]
    headers, err = _read_headers(ws)
    if err:
        return {}, [f"{sheet_name}: {err}"]

    missing = _REQUIRED_OBJ_HEADERS - set(headers)
    if missing:
        return {}, [f"{sheet_name} missing required columns: {missing}"]

    assets = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(_blank(c) for c in row):
            continue
        r = _row_map(headers, row)
        name = _str(r.get("object_id"))
        if not name:
            continue
        assets[name] = Asset(
            object_id     = name,
            locator_type  = _str(r.get("locator_type")).upper() or LOCATOR_IMAGE,
            resource_path = _str(r.get("resource_path")),
            threshold     = _float(r.get("smart_threshold"), 0.8),
            timeout       = _float(r.get("timeout"), 5.0),
        )
    return assets, []


def parse_action_logic(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return set(), [], []

    ws = wb[sheet_name]
    headers, err = _read_headers(ws)
    if err:
        return set(), [], [f"{sheet_name}: {err}"]

    missing = _REQUIRED_ACTION_HEADERS - set(headers)
    if missing:
        return set(), [], [f"{sheet_name} missing required columns: {missing}"]

    keywords = set()
    flows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(_blank(c) for c in row):
            continue
        r = _row_map(headers, row)
        logic_id = _str(r.get("logic_id"))
        name     = _str(r.get("action_name"))
        if not logic_id or not name:
            continue
        if logic_id.upper().startswith("ACT"):
            keywords.add(name.upper())
        else:
            flows.append(FlowDoc(
                logic_id    = logic_id,
                action_name = name,
                command     = _str(r.get("machine_command")),
                target_page = _str(r.get("target_page")),
            ))
    return keywords, flows, []


def parse_test_execution(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return {}, [f"Sheet '{sheet_name}' not found"]

    ws = wb[sheet_name]
    headers, err = _read_headers(ws)
    if err:
        return {}, [f"{sheet_name}: {err}"]

    missing = _REQUIRED_STEP_HEADERS - set(headers)
    if missing:
        return {}, [f"{sheet_name} missing required columns: {missing}"]

    suites = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(_blank(c) for c in row):
            continue
        r = _row_map(headers, row)
        suite = _str(r.get("suite_id"))
        if not suite:
            continue
        suites.setdefault(suite, []).append(Step(
            suite_id  = suite,
            step_no   = _int(r.get("step"), 0),
            action    = _str(r.get("action_keyword")).upper(),
            excel_row = row_num,
            target    = _str(r.get("target_id")),
            params    = _str(r.get("params")),
            expected  = _str(r.get("expected_result")),
        ))
    return suites, []


# --------------------------------------------------------------------------- #
# Validator                                                                   #
# --------------------------------------------------------------------------- #

def validate_assets(assets, base_dir):
    issues = []
    for name, asset in assets.items():
        if asset.locator_type != LOCATOR_IMAGE:
            continue
        path = asset.resource_path
        if not path or path.upper() == "NONE":
            continue
        full = path if os.path.isabs(path) else os.path.join(base_dir, path)
        if not os.path.isfile(full):
            issues.append(ValidationIssue(component=name, path=path))
            print(f"  [WARN] Asset not found on disk: {path}  (object: {name})")
    return issues


# --------------------------------------------------------------------------- #
# Action Registry (per-class)                                                 #
# --------------------------------------------------------------------------- #

def action(name: str):
    """Decorator that tags a method as the handler for an action keyword.

    Subclasses inherit parent's handlers via the _GenMeta metaclass; a child
    decorating the same name overrides the parent's handler.
    """
    def deco(fn: Callable) -> Callable:
        fn._action_name = name.upper()
        return fn
    return deco


class _GenMeta(type):
    """Collect every @action()-tagged method into a per-class _HANDLERS dict.

    Subclasses inherit parents' handlers first, then their own methods override
    by action-name. The resulting dict lives on the class itself.
    """
    def __new__(mcs, name, bases, ns):
        handlers = {}
        for base in reversed(bases):
            handlers.update(getattr(base, "_HANDLERS", {}))
        for v in ns.values():
            if callable(v) and hasattr(v, "_action_name"):
                handlers[v._action_name] = v
        ns["_HANDLERS"] = handlers
        return super().__new__(mcs, name, bases, ns)


# --------------------------------------------------------------------------- #
# Project registry                                                             #
# --------------------------------------------------------------------------- #

_PROJECT_REGISTRY: dict[str, type] = {}


def register_project(name: str):
    """Decorator: register a generator subclass under a project name.

    Usage::

        @register_project("mygame")
        class MyGameGenerator(AirtestGenerator):
            DEFAULT_APP_PACKAGE = "com.my.game"
    """
    def deco(cls):
        _PROJECT_REGISTRY[name.lower()] = cls
        return cls
    return deco


def _classname_from_name(name: str) -> str:
    """'my_game' -> 'MyGame'  (used by --init-project scaffold)."""
    return "".join(part.capitalize() for part in re.split(r"[_\-\s]+", name)) or "Project"


def _discover_projects(projects_dir: str = None) -> None:
    """Import all *.py files from projects/ so their @register_project decorators fire."""
    if projects_dir is None:
        projects_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "projects")
    if not os.path.isdir(projects_dir):
        return
    parent = os.path.dirname(projects_dir)
    if parent not in sys.path:
        sys.path.insert(0, parent)
    # When running as __main__, project files do `from excel_to_airtest import register_project`
    # which would load a *second* copy of this module and write to a different registry.
    # Register __main__ under the canonical name so they all share one registry.
    if __name__ == "__main__" and "excel_to_airtest" not in sys.modules:
        sys.modules["excel_to_airtest"] = sys.modules["__main__"]
    for fname in sorted(os.listdir(projects_dir)):
        if fname.endswith(".py") and not fname.startswith("_"):
            module_name = f"projects.{fname[:-3]}"
            if module_name not in sys.modules:
                importlib.import_module(module_name)


_SCAFFOLD_TEMPLATE = '''\
# projects/{name}.py  — generated scaffold
# Edit DEFAULT_APP_PACKAGE and IMPORTS, then run:
#   python excel_to_airtest.py --list-projects
# to confirm your project is discovered.

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from excel_to_airtest import AirtestGenerator, register_project, action


@register_project("{name}")
class {classname}Generator(AirtestGenerator):
    DEFAULT_APP_PACKAGE = "com.example.{name}"   # <- change this
    IMPORTS = "from airtest.core.api import *"    # <- change this
    MODULE_PROLOGUE = ""                           # <- page-object singletons (optional)

    # Override actions only if your framework differs from base Airtest.
    # All base actions are inherited automatically.

    # @action("TAP")
    # def handle_tap(self, step, ctx):
    #     asset, lines, issue = self._resolve_image(step, ctx)
    #     if asset is None: return lines, issue
    #     return [f"my_framework.tap({{asset.resource_path!r}})"], None

    # def wrap_main_body(self, step_lines, suite_id):
    #     return step_lines


if __name__ == "__main__":
    {classname}Generator.main()
'''


def _scaffold_project(name: str, projects_dir: str = None) -> str:
    """Write projects/<name>.py scaffold. Raises FileExistsError if already present."""
    if projects_dir is None:
        projects_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "projects")
    os.makedirs(projects_dir, exist_ok=True)
    out_path = os.path.join(projects_dir, f"{name}.py")
    if os.path.exists(out_path):
        raise FileExistsError(f"Project file already exists: {out_path}")
    classname = _classname_from_name(name)
    content = _SCAFFOLD_TEMPLATE.format(name=name, classname=classname)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(content)
    return out_path


# --------------------------------------------------------------------------- #
# Diagnostic hints                                                            #
# --------------------------------------------------------------------------- #

_HINT_MAP = {
    "MISSING_TARGET":           "fill the Target_ID cell for this step in Test_Execution sheet",
    "START_APP_NEEDS_PACKAGE":  "pass --app-package com.your.app on the CLI",
    "STOP_APP_NEEDS_PACKAGE":   "pass --app-package com.your.app on the CLI",
    "INVALID_PARAMS_JSON":      "Params must be valid JSON or a plain number",
    "SWIPE_NEEDS_PARAMS":       'provide JSON {"from":[x1,y1],"to":[x2,y2],"duration":0.5} in Params',
    "INVALID_SWIPE_PARAMS":     'Params must be valid JSON: {"from":[x1,y1],"to":[x2,y2],"duration":0.5}',
    "INVALID_SCROLL_DIRECTION": "use one of: up/down/left/right in the Params cell",
    "INVALID_SLEEP_PARAMS":     'provide seconds as a number (e.g. "2.5") or {"seconds": 2.5}',
    "READ_TEXT":                "READ_TEXT is a v2 feature — leave as TODO stub for now",
}


def _diagnostic_hint(reason: str) -> str:
    """Return an actionable fix hint for a generation issue reason string."""
    for key, hint in _HINT_MAP.items():
        if key in reason:
            return hint
    if "UNKNOWN_TARGET" in reason:
        target = reason.split("'")[1] if "'" in reason else reason
        return f"add row with Object_ID='{target}' to Object_Repository sheet"
    if "UNSUPPORTED_LOCATOR" in reason:
        return "OCR locators are v2 — change Locator_Type to IMAGE or leave as TODO stub"
    if "MISSING_RESOURCE_PATH" in reason:
        target = reason.split("'")[1] if "'" in reason else ""
        return f"fill Resource_Path for '{target}' in Object_Repository sheet"
    return "check the Excel cell for this step"


# --------------------------------------------------------------------------- #
# Generator base class - subclass to retarget for any project                 #
# --------------------------------------------------------------------------- #

class AirtestGenerator(metaclass=_GenMeta):
    """Father tool. Subclass and override hooks to retarget for any project."""

    # === Sheet names (subclass may rebind) ============================== #
    OBJECTS_SHEET = "Object_Repository"
    ACTIONS_SHEET = "Action_Logic"
    STEPS_SHEET   = "Test_Execution"

    # === Generated-script template hooks ================================ #
    # Source lines that go at the top of every generated .py file.
    IMPORTS = "from airtest.core.api import *\nfrom airtest.aircv import Template"

    # Lines at module scope BELOW imports and ABOVE def main(). Use for
    # singletons / page-object instantiation that every test shares.
    MODULE_PROLOGUE = ""

    # === CLI defaults (subclass may rebind) ============================= #
    DESCRIPTION         = "Convert AutomationRebase Excel to Airtest .air scripts (one per Suite_ID)."
    DEFAULT_APP_PACKAGE = ""
    DEFAULT_PLAN        = None

    # ------------------------------------------------------------------- #
    # Asset resolution helpers (subclass may reuse or override)           #
    # ------------------------------------------------------------------- #

    def _issue(self, step: Step, reason: str) -> GenerationIssue:
        return GenerationIssue(step.suite_id, step.step_no, step.excel_row, reason)

    def _todo(self, step: Step, reason: str):
        return [f"# TODO: {reason}"], self._issue(step, reason)

    def _resolve_image(self, step: Step, ctx: GenCtx):
        """Return (Asset, None, None) on success or (None, lines, issue) on failure."""
        if not step.target:
            l, i = self._todo(step, "MISSING_TARGET")
            return None, l, i
        asset = ctx.assets.get(step.target)
        if asset is None:
            l, i = self._todo(step, f"UNKNOWN_TARGET '{step.target}'")
            return None, l, i
        if asset.locator_type != LOCATOR_IMAGE:
            l, i = self._todo(step, f"UNSUPPORTED_LOCATOR '{asset.locator_type}' for '{step.target}'")
            return None, l, i
        if not asset.resource_path or asset.resource_path.upper() == "NONE":
            l, i = self._todo(step, f"MISSING_RESOURCE_PATH for '{step.target}'")
            return None, l, i
        return asset, None, None

    # ------------------------------------------------------------------- #
    # Default action handlers (bare Airtest)                              #
    # ------------------------------------------------------------------- #

    @action("TAP")
    def handle_tap(self, step, ctx):
        asset, lines, issue = self._resolve_image(step, ctx)
        if asset is None:
            return lines, issue
        return [f"touch(Template({asset.resource_path!r}, threshold={asset.threshold}))"], None

    @action("TOUCH")
    def handle_touch(self, step, ctx):
        return self.handle_tap(step, ctx)

    @action("CLICK")
    def handle_click(self, step, ctx):
        asset, lines, issue = self._resolve_image(step, ctx)
        if asset is None:
            return lines, issue
        return [
            "# deprecated: use TAP or TOUCH",
            f"touch(Template({asset.resource_path!r}, threshold={asset.threshold}))",
        ], None

    @action("WAIT_FOR")
    def handle_wait_for(self, step, ctx):
        asset, lines, issue = self._resolve_image(step, ctx)
        if asset is None:
            return lines, issue
        return [f"wait(Template({asset.resource_path!r}, threshold={asset.threshold}), timeout={asset.timeout})"], None

    @action("ASSERT_VISIBLE")
    def handle_assert_visible(self, step, ctx):
        asset, lines, issue = self._resolve_image(step, ctx)
        if asset is None:
            return lines, issue
        return [f"assert_exists(Template({asset.resource_path!r}, threshold={asset.threshold}), timeout={asset.timeout})"], None

    @action("START_APP")
    def handle_start_app(self, step, ctx):
        out = []
        if step.params:
            try:
                cfg = json.loads(step.params)
                out.append(f"# config: {cfg!r}")
            except json.JSONDecodeError:
                return [f"# TODO: INVALID_PARAMS_JSON: {step.params!r}"], self._issue(step, "INVALID_PARAMS_JSON")
        if not ctx.app_package:
            out.append("# TODO: START_APP_NEEDS_PACKAGE - pass via --app-package")
            return out, self._issue(step, "START_APP_NEEDS_PACKAGE")
        out.append(f"start_app({ctx.app_package!r})")
        return out, None

    @action("INPUT_TEXT")
    def handle_input_text(self, step, ctx):
        return [f"text({step.params!r})"], None

    @action("READ_TEXT")
    def handle_read_text(self, step, ctx):
        return self._todo(step, f"READ_TEXT not implemented (target='{step.target}')")

    @action("SWIPE")
    def handle_swipe(self, step, ctx):
        if not step.params:
            return self._todo(step, 'SWIPE_NEEDS_PARAMS: provide JSON {"from":[x1,y1],"to":[x2,y2],"duration":0.5}')
        try:
            cfg = json.loads(step.params)
            x1, y1 = cfg["from"]
            x2, y2 = cfg["to"]
            duration = cfg.get("duration", 0.5)
            return [f"swipe(({x1}, {y1}), ({x2}, {y2}), duration={duration})"], None
        except (json.JSONDecodeError, KeyError, TypeError, ValueError):
            return self._todo(step, f"INVALID_SWIPE_PARAMS: {step.params!r}")

    @action("SCROLL")
    def handle_scroll(self, step, ctx):
        raw = (step.params.strip() if step.params else "") or "up"
        try:
            direction = json.loads(raw)["direction"].lower()
        except (json.JSONDecodeError, KeyError, TypeError, AttributeError):
            direction = raw.lower()
        if direction not in _SCROLL_PRESETS:
            return self._todo(step, f"INVALID_SCROLL_DIRECTION: '{direction}' — use up/down/left/right")
        frm, to = _SCROLL_PRESETS[direction]
        return [
            "w, h = G.DEVICE.get_current_resolution()",
            f"swipe({frm}, {to})",
        ], None

    @action("LONG_PRESS")
    def handle_long_press(self, step, ctx):
        asset, lines, issue = self._resolve_image(step, ctx)
        if asset is None:
            return lines, issue
        return [f"long_click(Template({asset.resource_path!r}, threshold={asset.threshold}))"], None

    @action("SLEEP")
    def handle_sleep(self, step, ctx):
        raw = step.params.strip() if step.params else ""
        try:
            if raw.startswith("{"):
                seconds = float(json.loads(raw)["seconds"])
            else:
                seconds = float(raw)
            return [f"sleep({seconds})"], None
        except (json.JSONDecodeError, KeyError, ValueError, TypeError):
            return self._todo(step, f"INVALID_SLEEP_PARAMS: {step.params!r} — provide seconds as number or {{\"seconds\": 2.5}}")

    @action("BACK")
    def handle_back(self, step, ctx):
        return ['keyevent("BACK")'], None

    @action("HOME")
    def handle_home(self, step, ctx):
        return ['keyevent("HOME")'], None

    @action("SNAPSHOT")
    def handle_snapshot(self, step, ctx):
        filename = (step.params.strip() if step.params else "") or step.suite_id.lower()
        if not filename.endswith(".png"):
            filename += ".png"
        return [f"snapshot(filename={filename!r})"], None

    @action("STOP_APP")
    def handle_stop_app(self, step, ctx):
        if not ctx.app_package:
            return self._todo(step, "STOP_APP_NEEDS_PACKAGE — pass via --app-package")
        return [f'stop_app("{ctx.app_package}")'], None

    # ------------------------------------------------------------------- #
    # Generation                                                          #
    # ------------------------------------------------------------------- #

    def step_label(self, step: Step) -> str:
        return f"# step {step.step_no}" + (f" - {step.expected}" if step.expected else "")

    def _generate_step(self, step: Step, ctx: GenCtx):
        label = self.step_label(step)
        handler = self._HANDLERS.get(step.action)
        if handler is None:
            return [label, f"# UNSUPPORTED_ACTION: {step.action!r}"], self._issue(step, f"UNSUPPORTED_ACTION '{step.action}'")
        lines, issue = handler(self, step, ctx)
        body = [l for l in lines if l]
        return [label] + body, issue

    def wrap_main_body(self, step_lines, suite_id):
        """Wrap the def main() body. Default: identity.

        Override to inject try/except,
        snapshot-on-fail, teardown, etc.

        Receives a flat list of unindented lines (one statement / comment per
        list entry). Must return a flat list of unindented lines; the framework
        will indent the result by 4 spaces (one level inside def main()).
        Internal indentation inside returned lines is preserved.
        """
        return step_lines

    def generate_suite_script(self, steps, ctx, source_name, suite_id):
        body_lines = []
        issues = []
        for step in steps:
            lines, issue = self._generate_step(step, ctx)
            body_lines.extend(lines)
            if issue:
                issues.append(issue)

        wrapped = self.wrap_main_body(body_lines, suite_id)

        # Function body must contain at least one statement. Comment-only
        # bodies (e.g. suite of TODO stubs) are a syntax error - emit pass.
        if not any(l and not l.lstrip().startswith("#") for l in wrapped):
            wrapped.append("pass")

        indent = "    "
        body = ("\n" + indent).join(wrapped) if wrapped else "pass"

        prologue = self.MODULE_PROLOGUE.strip()
        prologue_block = ("\n\n" + prologue) if prologue else ""

        return (
            f"# Auto-generated by excel_to_airtest.py ({type(self).__name__})\n"
            f"# Source: {source_name}\n"
            f"# Suite : {suite_id}\n"
            f"{self.IMPORTS}"
            f"{prologue_block}\n"
            f"\n\n"
            f"def main():\n"
            f"    {body}\n"
            f"\n\n"
            f'if __name__ == "__main__":\n'
            f"    main()\n"
        ), issues

    # ------------------------------------------------------------------- #
    # Writer                                                              #
    # ------------------------------------------------------------------- #

    def write_suite(self, script, plan_dir, suite_id):
        safe = _safe_name(suite_id)
        air_dir = os.path.join(plan_dir, f"{safe}.air")
        os.makedirs(air_dir, exist_ok=True)
        out_file = os.path.join(air_dir, f"{safe}.py")
        with open(out_file, "w", encoding="utf-8") as f:
            f.write(script)
        return out_file

    def write_report(self, output_dir, source, written, gen_issues, val_issues, flows):
        lines = [
            "=== Generation Report ===",
            f"Generator : {type(self).__name__}",
            f"Source    : {source}",
            "",
            f"--- Generated Suites ({len(written)}) ---",
        ]
        for suite, path in written:
            lines.append(f"  {suite}  ->  {path}")
        if not written:
            lines.append("  (none)")

        lines += ["", f"--- Generation Issues ({len(gen_issues)}) ---"]
        if gen_issues:
            for i in gen_issues:
                lines.append(f"  row {i.excel_row:>4} | suite {i.suite_id} | step {i.step_no}: {i.reason}")
        else:
            lines.append("  (none)")

        lines += ["", f"--- Asset Validation - file not found ({len(val_issues)}) ---"]
        if val_issues:
            for v in val_issues:
                lines.append(f"  {v.component}: {v.path}")
        else:
            lines.append("  (none)")

        lines += ["", f"--- Action_Logic Flows ({len(flows)}) - reference only, not generated ---"]
        if flows:
            for fl in flows:
                lines.append(f"  {fl.logic_id} | {fl.action_name} ({fl.target_page})")
        else:
            lines.append("  (none)")

        report_path = os.path.join(output_dir, "generation_report.txt")
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")
        return report_path

    # ------------------------------------------------------------------- #
    # CLI                                                                 #
    # ------------------------------------------------------------------- #

    def add_arguments(self, parser):
        """Subclasses may extend by calling super().add_arguments(parser) first."""
        parser.add_argument("excel_file", nargs="?", default=None,
                            help="Path to Excel (.xlsx) file")
        parser.add_argument("--output",        default="./output",
                            help="Output root directory")
        parser.add_argument("--plan",          default=self.DEFAULT_PLAN,
                            help="Subfolder under output (default: excel filename stem)")
        parser.add_argument("--objects-sheet", default=self.OBJECTS_SHEET,
                            help="Object repository sheet name")
        parser.add_argument("--actions-sheet", default=self.ACTIONS_SHEET,
                            help="Action logic sheet name")
        parser.add_argument("--steps-sheet",   default=self.STEPS_SHEET,
                            help="Step sheet name")
        parser.add_argument("--app-package",   default=self.DEFAULT_APP_PACKAGE,
                            help="App package id used by START_APP / STOP_APP")
        parser.add_argument("--report",        action="store_true",
                            help="Write generation_report.txt")
        parser.add_argument("--project",       default=None,
                            help="Project name to use (from projects/ directory)")
        parser.add_argument("--list-projects", action="store_true",
                            help="List all discovered projects and exit")
        parser.add_argument("--init-project",  metavar="NAME",
                            help="Scaffold projects/<NAME>.py and exit")
        parser.add_argument("--init-excel",    action="store_true",
                            help="Generate template.xlsx in current directory and exit")

    def run(self, args):
        if not os.path.isfile(args.excel_file):
            sys.exit(f"ERROR: File not found: {args.excel_file}")

        plan = args.plan or os.path.splitext(os.path.basename(args.excel_file))[0]
        base_dir = os.path.dirname(os.path.abspath(args.excel_file))
        plan_dir = os.path.join(args.output, plan)
        os.makedirs(plan_dir, exist_ok=True)

        wb = openpyxl.load_workbook(args.excel_file, data_only=True)

        print(f"[parse] {args.objects_sheet}...")
        assets, obj_errors = parse_object_repository(wb, args.objects_sheet)
        for e in obj_errors: print(f"  [ERROR] {e}")
        if obj_errors: sys.exit(1)
        print(f"        {len(assets)} objects loaded")

        print(f"[parse] {args.actions_sheet}...")
        keywords, flows, act_errors = parse_action_logic(wb, args.actions_sheet)
        for e in act_errors: print(f"  [ERROR] {e}")
        print(f"        {len(keywords)} action keywords, {len(flows)} flow descriptions")

        print("[validate] Asset paths on disk...")
        val_issues = validate_assets(assets, base_dir)
        print(f"           {len(val_issues)} missing file(s)")

        print(f"[parse] {args.steps_sheet}...")
        suites, step_errors = parse_test_execution(wb, args.steps_sheet)
        for e in step_errors: print(f"  [ERROR] {e}")
        if step_errors: sys.exit(1)
        print(f"        {len(suites)} suite(s), {sum(len(v) for v in suites.values())} step(s)")

        ctx = GenCtx(assets=assets, app_package=args.app_package)
        source_name = os.path.basename(args.excel_file)

        print(f"[generate] Building scripts ({type(self).__name__})...")
        written = []
        gen_issues = []
        for suite_id, steps in suites.items():
            script, issues = self.generate_suite_script(steps, ctx, source_name, suite_id)
            out_path = self.write_suite(script, plan_dir, suite_id)
            written.append((suite_id, out_path))
            gen_issues.extend(issues)
            print(f"  [write] {suite_id} -> {out_path}  ({len(issues)} issue(s))")

        if args.report:
            report = self.write_report(args.output, args.excel_file, written, gen_issues, val_issues, flows)
            print(f"[report] {report}")

        total = len(gen_issues) + len(val_issues)
        if gen_issues or val_issues:
            print(f"\n[issues] {len(gen_issues) + len(val_issues)} problem(s) found:")
            for i in gen_issues:
                hint = _diagnostic_hint(i.reason)
                print(f"  row {i.excel_row:>4} | {i.suite_id} | step {i.step_no} | {i.reason}")
                print(f"          -> fix: {hint}")
            for v in val_issues:
                print(f"  asset    | {v.component}: {v.path}")
                print(f"          -> fix: copy or create this image file at the listed path")
        print(f"\n{'OK - no issues' if total == 0 else f'DONE - {len(gen_issues)} generation issue(s), {len(val_issues)} missing asset(s)'}")

    @classmethod
    def main(cls):
        _discover_projects()
        parser = argparse.ArgumentParser(description=cls.DESCRIPTION)
        gen = cls()
        gen.add_arguments(parser)
        args = parser.parse_args()

        if args.list_projects:
            if not _PROJECT_REGISTRY:
                print("No projects registered. Add a file to projects/ or run --init-project <name>.")
            for name, klass in sorted(_PROJECT_REGISTRY.items()):
                src = sys.modules.get(klass.__module__, None)
                src_file = getattr(src, "__file__", "?") if src else "?"
                print(f"  {name:<16} {src_file:<40} {klass.DEFAULT_APP_PACKAGE}")
            return

        if args.init_project:
            try:
                path = _scaffold_project(args.init_project)
                print(f"[init-project] Created: {path}")
                print(f"               Edit DEFAULT_APP_PACKAGE and IMPORTS, then run --list-projects.")
            except FileExistsError as e:
                sys.exit(f"ERROR: {e}")
            return

        if args.init_excel:
            from templates.excel_template import generate_template
            out = generate_template("template.xlsx")
            print(f"[init-excel] Template written to: {out}")
            return

        # Select project class
        gen_cls = cls
        if args.project:
            key = args.project.lower()
            if key not in _PROJECT_REGISTRY:
                known = ", ".join(sorted(_PROJECT_REGISTRY)) or "(none)"
                sys.exit(f"ERROR: Unknown project '{args.project}'. Known: {known}. Run --list-projects.")
            gen_cls = _PROJECT_REGISTRY[key]

        if not args.excel_file:
            parser.print_help()
            sys.exit("\nERROR: excel_file is required.")

        gen_cls().run(args)


# --------------------------------------------------------------------------- #
# Backward-compat module-level surface                                        #
# --------------------------------------------------------------------------- #
#
# Older code (and tests) call these by module-level names. They delegate to a
# default AirtestGenerator() instance so existing imports keep working.

_default_generator = AirtestGenerator()

_HANDLERS = AirtestGenerator._HANDLERS


def _issue(step, reason):
    return _default_generator._issue(step, reason)


def _todo(step, reason):
    return _default_generator._todo(step, reason)


def _resolve_image(step, ctx):
    return _default_generator._resolve_image(step, ctx)


def generate_suite_script(steps, ctx, source_name, suite_id):
    return _default_generator.generate_suite_script(steps, ctx, source_name, suite_id)


def main():
    AirtestGenerator.main()


if __name__ == "__main__":
    main()
