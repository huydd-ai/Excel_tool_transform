"""
rules_to_suites.py — Convert Manual_Testing_Guidelines_Rules.xlsx into .air scripts.

Reads the rules-based Excel schema (General Guidelines, Feature Rules, Edge Cases,
Release Checklist) and maps each rule/check to a test suite using the existing
AirtestGenerator pipeline.

Usage:
    python rules_to_suites.py Manual_Testing_Guidelines_Rules.xlsx --project pixon
    python rules_to_suites.py Manual_Testing_Guidelines_Rules.xlsx --output ./output --app-package com.demo
"""

import argparse
import json
import os
import sys
from dataclasses import dataclass, field

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

from excel_to_airtest import (
    AirtestGenerator,
    Asset,
    GenCtx,
    Step,
    _discover_projects,
    _PROJECT_REGISTRY,
)
from models import AirtestError
from parsers import _str

# --------------------------------------------------------------------------- #
# Data models for the rules document                                          #
# --------------------------------------------------------------------------- #


@dataclass
class FeatureRule:
    feature: str
    logic_item: str
    condition: str
    expected: str


@dataclass
class EdgeCase:
    edge_id: str
    scenario: str
    condition: str
    recovery: str


@dataclass
class ChecklistItem:
    check_id: str
    description: str


@dataclass
class RulesDoc:
    guidelines: list = field(default_factory=list)
    feature_rules: list = field(default_factory=list)
    edge_cases: list = field(default_factory=list)
    checklist: list = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Parser: read the manual-testing-guidelines Excel                            #
# --------------------------------------------------------------------------- #


def _header_index(ws) -> dict[str, int]:
    """Map normalised (lowercased, stripped) header text -> column index."""
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if hdr is None:
        return {}
    return {_str(h).lower(): i for i, h in enumerate(hdr) if not _blank_cell(h)}


def _blank_cell(val) -> bool:
    return val is None or (isinstance(val, str) and not val.strip())


def _col(row: tuple, idx: int | None, fallback: int) -> str:
    """Read a cell by header index, falling back to a fixed position.

    Header-mapping wins when the column is present (robust to reorder);
    positional fallback keeps working when header text drifts.
    """
    i = idx if idx is not None else fallback
    return _str(row[i]) if 0 <= i < len(row) else ""


def read_rules_excel(path: str) -> RulesDoc:
    wb = openpyxl.load_workbook(path, data_only=True)
    doc = RulesDoc()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        hdr = _header_index(ws)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            continue

        if sheet_name == "General Guidelines":
            doc.guidelines = [
                (r[0], r[1], r[2]) for r in rows if any(c is not None for c in r)
            ]

        elif sheet_name == "Feature Rules":
            for r in rows:
                if not any(c is not None for c in r):
                    continue
                doc.feature_rules.append(
                    FeatureRule(
                        feature=_col(r, hdr.get("feature"), 0),
                        logic_item=_col(r, hdr.get("logic item"), 1),
                        condition=_col(r, hdr.get("rule / condition"), 2),
                        expected=_col(r, hdr.get("expected behavior"), 3),
                    )
                )

        elif sheet_name == "Edge Cases":
            for r in rows:
                if not any(c is not None for c in r):
                    continue
                doc.edge_cases.append(
                    EdgeCase(
                        edge_id=_col(r, hdr.get("id"), 0),
                        scenario=_col(r, hdr.get("scenario"), 1),
                        condition=_col(r, hdr.get("condition"), 2),
                        recovery=_col(r, hdr.get("required handling (recovery)"), 3),
                    )
                )

        elif sheet_name == "Release Checklist":
            for r in rows:
                if not any(c is not None for c in r):
                    continue
                doc.checklist.append(
                    ChecklistItem(
                        check_id=_col(r, hdr.get("check id"), 0),
                        description=_col(r, hdr.get("verify item"), 1),
                    )
                )

    return doc


# --------------------------------------------------------------------------- #
# Suite definition                                                            #
# --------------------------------------------------------------------------- #


@dataclass
class SuiteDef:
    suite_id: str
    steps: list
    assets: dict = field(default_factory=dict)
    desc: str = ""


# --------------------------------------------------------------------------- #
# Object_Repository entries shared across rules suites                        #
# --------------------------------------------------------------------------- #

# Placeholder assets — testers must capture actual screenshots.
_BASE_ASSETS: dict[str, Asset] = {
    "heart_icon": Asset(
        "heart_icon", resource_path="./assets/heart_icon.png", timeout=10
    ),
    "btn_play": Asset("btn_play", resource_path="./assets/btn_play.png", timeout=10),
    "confirm_popup": Asset(
        "confirm_popup", resource_path="./assets/confirm_popup.png", timeout=10
    ),
    "btn_confirm": Asset(
        "btn_confirm", resource_path="./assets/btn_confirm.png", timeout=10
    ),
    "exp_icon": Asset("exp_icon", resource_path="./assets/exp_icon.png", timeout=10),
    "heart_max_8": Asset(
        "heart_max_8", resource_path="./assets/heart_max_8.png", timeout=10
    ),
    "timer_display": Asset(
        "timer_display", resource_path="./assets/timer_display.png", timeout=10
    ),
    "cooldown_screen": Asset(
        "cooldown_screen", resource_path="./assets/cooldown_screen.png", timeout=10
    ),
    "exp_bar": Asset("exp_bar", resource_path="./assets/exp_bar.png", timeout=10),
    "unlimited_timer": Asset(
        "unlimited_timer", resource_path="./assets/unlimited_timer.png", timeout=10
    ),
    "heart_unchanged": Asset(
        "heart_unchanged", resource_path="./assets/heart_unchanged.png", timeout=10
    ),
    "heart_4": Asset("heart_4", resource_path="./assets/heart_4.png", timeout=10),
    "heart_3": Asset("heart_3", resource_path="./assets/heart_3.png", timeout=10),
}


# --------------------------------------------------------------------------- #
# Step builders                                                               #
# --------------------------------------------------------------------------- #

# Step builders all emit step_no=0; resolve_suites() reindexes 1..N per suite
# via _reindex_steps(), so the placeholder value is never used downstream.


def _start_app(cfg: dict | None = None) -> Step:
    return Step(
        suite_id="",
        step_no=0,
        action="START_APP",
        params=json.dumps(cfg) if cfg else "",
        expected="App launches with test config",
    )


def _tap(target: str, desc: str = "") -> Step:
    return Step(suite_id="", step_no=0, action="TAP", target=target, expected=desc)


def _wait(target: str, desc: str = "") -> Step:
    return Step(suite_id="", step_no=0, action="WAIT_FOR", target=target, expected=desc)


def _assert(target: str, desc: str = "") -> Step:
    return Step(
        suite_id="", step_no=0, action="ASSERT_VISIBLE", target=target, expected=desc
    )


def _sleep(secs: str) -> Step:
    return Step(suite_id="", step_no=0, action="SLEEP", params=secs)


def _snapshot(name: str) -> Step:
    return Step(suite_id="", step_no=0, action="SNAPSHOT", params=name)


def _todo(reason: str) -> Step:
    return Step(
        suite_id="", step_no=0, action="SLEEP", params="0", expected=f"# TODO: {reason}"
    )


def _todo_step(reason: str) -> Step:
    return Step(
        suite_id="",
        step_no=0,
        action="SLEEP",
        params="0",
        expected=f"{reason}",
        excel_row=0,
        target="",
    )


# --------------------------------------------------------------------------- #
# Rule mapping — each feature rule, edge case, checklist item -> SuiteDef     #
# --------------------------------------------------------------------------- #


def _suite_id(name: str) -> str:
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in name)
    return safe.upper().strip("_")


def _assets_from_targets(steps: list[Step]) -> dict[str, Asset]:
    needed = {}
    for s in steps:
        target = s.target.strip() if s.target else ""
        if target and target in _BASE_ASSETS:
            needed[target] = _BASE_ASSETS[target]
    return needed


# Each builder is zero-arg and returns (suite_id, steps). Per-suite `assets` are
# auto-derived from step targets in resolve_suites() (see _assets_from_targets),
# and `desc` is filled centrally from the originating rule. This keeps the maps as
# pure step-data with no repeated assets/desc boilerplate.

# --- Feature Rules ---

FEATURE_RULE_MAP = {
    "Passive Regen": lambda: (
        "HEART_PassiveRegen",
        [
            _start_app({"heart": 4}),
            _todo_step("MANUAL_TIMER_CHECK — 20 min regen; skip in automation"),
            _todo_step("VERIFY timer persists after app restart (offline)"),
        ],
    ),
    "Consumption": lambda: (
        "HEART_Consumption",
        [
            _start_app({"heart": 5}),
            _tap("btn_play", "Start level"),
            _wait("confirm_popup", "Warning popup appears"),
            _tap("btn_confirm", "Confirm heart deduction"),
            _assert("heart_4", "Heart deducted from 5→4"),
        ],
    ),
    "Level Win": lambda: (
        "HEART_LevelWin",
        [
            _start_app({"heart": 3}),
            _todo_step("MANUAL_CHECK — play and win a level"),
            _assert("heart_3", "No heart deducted on win"),
        ],
    ),
    "Unlimited Heart": lambda: (
        "HEART_Unlimited",
        [
            _start_app({"heart": 0, "unlimited": True}),
            _assert("unlimited_timer", "Expiry countdown visible"),
            _todo_step("VERIFY no deduction when unlimited active"),
        ],
    ),
    "EXP Progression": lambda: (
        "ROYALPASS_EXP",
        [
            _start_app({"exp": 0}),
            _todo_step("MANUAL_CHECK — win a level and verify +1 EXP"),
            _assert("exp_icon", "EXP icon visible"),
            _todo_step("MANUAL_CHECK — lose level and verify no EXP change"),
        ],
    ),
    "Max Lives Buff": lambda: (
        "ROYALPASS_MaxLives",
        [
            _start_app({"pass": "gold"}),
            _assert("heart_max_8", "Max hearts increased to 8"),
        ],
    ),
    "Tier Unlocks": lambda: (
        "ROYALPASS_TierUnlock",
        [
            _start_app({"exp": "threshold"}),
            _todo_step("MANUAL_CHECK — verify tier unlock UI"),
            _todo_step("Ultimate Pass — verify +2 tiers ahead"),
        ],
    ),
    "Streak Logic": lambda: (
        "LAVAQUEST_Streak",
        [
            _start_app({"streak": 0}),
            _todo_step("MANUAL_CHECK — win 7 consecutive levels"),
            _todo_step("MANUAL_CHECK — loss/quit resets streak + 10 min cooldown"),
        ],
    ),
    "Reward Pool": lambda: (
        "LAVAQUEST_Reward",
        [
            _todo_step("MANUAL_CHECK — reward math: 5000 / (BOTs + 1)"),
        ],
    ),
}

# --- Edge Cases ---

EDGE_CASE_MAP = {
    "HEART-EDG-01": lambda: (
        "EDGE_HEART_ClockManipulation",
        [
            _todo_step("SERVER_VALIDATION — clock change, server time must override"),
        ],
    ),
    "HEART-EDG-02": lambda: (
        "EDGE_HEART_AppKill",
        [
            _start_app({"heart": 5}),
            _tap("btn_play", "Start level"),
            _sleep("2"),
            _todo_step("MANUAL_CHECK — force close app"),
            _start_app(),
            _assert("heart_icon", "Heart deducted after kill"),
        ],
    ),
    "RPS-EDG-03": lambda: (
        "EDGE_RPS_SimultaneousClaim",
        [
            _todo_step("SERVER_VALIDATION — idempotency check, first claim succeeds"),
        ],
    ),
    "LVQ-EDG-01": lambda: (
        "EDGE_LVQ_EventExpired",
        [
            _todo_step(
                "SERVER_VALIDATION — event expires mid-level, win allows advance"
            ),
            _todo_step("Show 'Event Over' after level ends"),
        ],
    ),
    "LVQ-EDG-02": lambda: (
        "EDGE_LVQ_BypassCooldown",
        [
            _todo_step(
                "SERVER_VALIDATION — clock change, server blocks entry for 10 min"
            ),
        ],
    ),
}

# --- Release Checklist ---

CHECKLIST_MAP = {
    "CHK-01": lambda: (
        "CHECK_HeartNoDeductOnWin",
        [
            _todo_step("MANUAL_CHECK — verify hearts not deducted on win"),
        ],
    ),
    "CHK-02": lambda: (
        "CHECK_GoldPassMax8",
        [
            _start_app({"pass": "gold"}),
            _assert("heart_max_8", "Gold Pass updates Heart Max to 8"),
        ],
    ),
    "CHK-03": lambda: (
        "CHECK_LavaStreakReset",
        [
            _todo_step("MANUAL_CHECK — streak resets on manual restart"),
        ],
    ),
    "CHK-04": lambda: (
        "CHECK_TimerFormat",
        [
            _todo_step("MANUAL_CHECK — timer format HH:MM:SS"),
        ],
    ),
    "CHK-05": lambda: (
        "CHECK_BuyHeartsOfflineSync",
        [
            _todo_step("MANUAL_CHECK — buy +5 hearts works offline/online sync"),
        ],
    ),
    "CHK-06": lambda: (
        "CHECK_RoyalPassEXPUI",
        [
            _start_app({"exp": 0}),
            _assert("exp_bar", "EXP updates in UI after level win"),
            _todo_step("MANUAL_CHECK — win level and verify EXP bar"),
        ],
    ),
    "CHK-07": lambda: (
        "CHECK_CooldownScreen",
        [
            _assert("cooldown_screen", "Cooldown screen visible for 10 mins"),
            _todo_step("MANUAL_CHECK — verify 10 min cooldown after Lava Quest loss"),
        ],
    ),
}


# --------------------------------------------------------------------------- #
# Resolve all suite defs from a RulesDoc                                      #
# --------------------------------------------------------------------------- #


def _make_suite(suite_id: str, steps: list[Step], desc: str) -> SuiteDef:
    """Assemble a SuiteDef: derive assets from step targets, then reindex steps."""
    sd = SuiteDef(suite_id=suite_id, steps=steps, desc=desc)
    sd.assets = _assets_from_targets(sd.steps)
    _reindex_steps(sd.steps, sd.suite_id)
    return sd


def resolve_suites(doc: RulesDoc) -> list[SuiteDef]:
    suites = []

    for rule in doc.feature_rules:
        desc = f"{rule.condition} | {rule.expected}"
        builder = FEATURE_RULE_MAP.get(rule.logic_item)
        if builder:
            suite_id, steps = builder()
        else:
            raise AirtestError(
                f"Unmapped rule logic item: '{rule.logic_item}' in feature '{rule.feature}'"
            )
        suites.append(_make_suite(suite_id, steps, desc))

    for edge in doc.edge_cases:
        desc = f"{edge.condition} | {edge.recovery}"
        builder = EDGE_CASE_MAP.get(edge.edge_id)
        if builder:
            suite_id, steps = builder()
        else:
            raise AirtestError(
                f"Unmapped edge case: '{edge.edge_id}' ({edge.scenario})"
            )
        suites.append(_make_suite(suite_id, steps, desc))

    for check in doc.checklist:
        builder = CHECKLIST_MAP.get(check.check_id)
        if builder:
            suite_id, steps = builder()
        else:
            raise AirtestError(
                f"Unmapped checklist item: '{check.check_id}' ({check.description})"
            )
        suites.append(_make_suite(suite_id, steps, check.description))

    return suites


def _reindex_steps(steps: list[Step], suite_id: str) -> list[Step]:
    for i, s in enumerate(steps, start=1):
        s.suite_id = suite_id
        s.step_no = i
    return steps


# --------------------------------------------------------------------------- #
# Resolve assets: merge per-suite assets with base assets                     #
# --------------------------------------------------------------------------- #


def _resolve_assets(suite_defs: list[SuiteDef]) -> dict[str, Asset]:
    combined = {}
    for sd in suite_defs:
        combined.update(sd.assets)
    return combined


# --------------------------------------------------------------------------- #
# Generate .air scripts for all suites                                        #
# --------------------------------------------------------------------------- #


def generate_all(
    suite_defs: list[SuiteDef],
    gen: AirtestGenerator,
    output_dir: str,
    plan: str,
    app_package: str,
    report: bool = False,
):
    assets = _resolve_assets(suite_defs)
    ctx = GenCtx(assets=assets, app_package=app_package)
    source_name = "Manual_Testing_Guidelines_Rules.xlsx"
    plan_dir = os.path.join(output_dir, plan)
    os.makedirs(plan_dir, exist_ok=True)

    written = []
    gen_issues = []
    for sd in suite_defs:
        script, issues = gen.generate_suite_script(
            sd.steps, ctx, source_name, sd.suite_id
        )
        out_path = gen.write_suite(script, plan_dir, sd.suite_id)
        written.append((sd.suite_id, out_path))
        gen_issues.extend(issues)

    if report:
        from excel_to_airtest import _diagnostic_hint

        lines = [
            "=== Generation Report (rules_to_suites) ===",
            f"Generator : {type(gen).__name__}",
            f"Source    : {source_name}",
            "",
            f"--- Generated Suites ({len(written)}) ---",
        ]
        for sid, path in written:
            lines.append(f"  {sid}  ->  {path}")
        lines += ["", f"--- Issues ({len(gen_issues)}) ---"]
        for i in gen_issues:
            hint = _diagnostic_hint(i.reason)
            lines.append(f"  {i.suite_id} | step {i.step_no} | {i.reason}")
            lines.append(f"          -> fix: {hint}")
        report_path = os.path.join(output_dir, "generation_report_rules.txt")
        try:
            with open(report_path, "w", encoding="utf-8") as f:
                f.write("\n".join(lines) + "\n")
        except OSError as exc:
            raise AirtestError(
                f"Failed to write report to {report_path}: {exc}"
            ) from exc
        print(f"[report] {report_path}")

    return written, gen_issues


# --------------------------------------------------------------------------- #
# CLI                                                                         #
# --------------------------------------------------------------------------- #


def main():
    _discover_projects()
    parser = argparse.ArgumentParser(
        description="Convert Manual_Testing_Guidelines_Rules.xlsx into .air scripts."
    )
    parser.add_argument(
        "excel_file", nargs="?", default=None, help="Path to rules Excel file"
    )
    parser.add_argument("--output", default="./output", help="Output root directory")
    parser.add_argument("--plan", default="rules", help="Subfolder under output")
    parser.add_argument("--app-package", default="", help="App package for START_APP")
    parser.add_argument("--project", default=None, help="Project name from projects/")
    parser.add_argument("--report", action="store_true", help="Write generation report")
    parser.add_argument(
        "--list-projects", action="store_true", help="List registered projects"
    )

    args = parser.parse_args()

    if args.list_projects:
        for name, klass in sorted(_PROJECT_REGISTRY.items()):
            src = sys.modules.get(klass.__module__, None)
            src_file = getattr(src, "__file__", "?") if src else "?"
            print(f"  {name:<16} {src_file:<40} {klass.DEFAULT_APP_PACKAGE}")
        return

    if not os.path.isfile(args.excel_file):
        sys.exit(f"ERROR: File not found: {args.excel_file}")

    print(f"[read] {args.excel_file}...")
    doc = read_rules_excel(args.excel_file)
    print(
        f"       {len(doc.feature_rules)} feature rules, "
        f"{len(doc.edge_cases)} edge cases, "
        f"{len(doc.checklist)} checklist items"
    )

    print("[resolve] Mapping rules to suites...")
    suite_defs = resolve_suites(doc)
    print(f"          {len(suite_defs)} suite(s)")

    gen_cls = AirtestGenerator
    if args.project:
        key = args.project.lower()
        if key not in _PROJECT_REGISTRY:
            known = ", ".join(sorted(_PROJECT_REGISTRY)) or "(none)"
            sys.exit(f"ERROR: Unknown project '{args.project}'. Known: {known}.")
        gen_cls = _PROJECT_REGISTRY[key]

    gen = gen_cls()
    app_pkg = args.app_package or gen_cls.DEFAULT_APP_PACKAGE
    print(f"[generate] Building scripts ({type(gen).__name__})...")
    written, gen_issues = generate_all(
        suite_defs,
        gen,
        output_dir=args.output,
        plan=args.plan,
        app_package=app_pkg,
        report=args.report,
    )

    for sid, path in written:
        print(f"  [write] {sid} -> {path}")

    total = len(gen_issues)
    if total:
        from excel_to_airtest import _diagnostic_hint

        print(f"\n[issues] {total} problem(s):")
        for i in gen_issues:
            hint = _diagnostic_hint(i.reason)
            print(f"  {i.suite_id} | step {i.step_no} | {i.reason}")
            print(f"          -> fix: {hint}")

    print(f"\n{'OK - no issues' if total == 0 else f'DONE - {total} issue(s)'}")


if __name__ == "__main__":
    main()
