"""Shared fixtures for excel_to_airtest tests.

Puts the project root on sys.path so tests can `import excel_to_airtest`
without an editable install.
"""

import os
import sys

import openpyxl
import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)


_OBJ_HEADERS = [
    "Object_ID",
    "Locator_Type",
    "Resource_Path",
    "Smart_Threshold",
    "Timeout",
]
_ACT_HEADERS = ["Logic_ID", "Action_Name", "Machine_Command", "Target_Page"]
_STEP_HEADERS = [
    "Suite_ID",
    "Step",
    "Action_Keyword",
    "Target_ID",
    "Params",
    "Expected_Result",
]


def _build_wb(
    objects=None,
    actions=None,
    steps=None,
    obj_headers=None,
    act_headers=None,
    step_headers=None,
):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    if objects is not None:
        ws = wb.create_sheet("Object_Repository")
        ws.append(obj_headers or _OBJ_HEADERS)
        for row in objects:
            ws.append(row)
    if actions is not None:
        ws = wb.create_sheet("Action_Logic")
        ws.append(act_headers or _ACT_HEADERS)
        for row in actions:
            ws.append(row)
    if steps is not None:
        ws = wb.create_sheet("Test_Execution")
        ws.append(step_headers or _STEP_HEADERS)
        for row in steps:
            ws.append(row)
    return wb


@pytest.fixture
def make_wb():
    """Return a Workbook builder. Pass objects=/actions=/steps= as lists of row tuples."""
    return _build_wb


@pytest.fixture
def make_wb_file(tmp_path, make_wb):
    """Return a builder that writes the Workbook to a temp file and returns its Path."""

    def _builder(**kw):
        wb = make_wb(**kw)
        path = tmp_path / "wb.xlsx"
        wb.save(path)
        return path

    return _builder
