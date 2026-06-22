"""Tests for rules_to_suites.py — parser, mapping, generation, inject_config."""

import os
import py_compile
import sys

import openpyxl
import pytest

from excel_to_airtest import AirtestGenerator, _HANDLERS
from models import GenCtx, Step
from rules_to_suites import (
    read_rules_excel,
    resolve_suites,
    generate_all,
    RulesDoc,
    FeatureRule,
    EdgeCase,
    ChecklistItem,
)
from models import AirtestError

# --------------------------------------------------------------------------- #
# Fixtures: build a minimal rules Excel in-memory                             #
# --------------------------------------------------------------------------- #


@pytest.fixture
def rules_wb_file(tmp_path):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "General Guidelines"
    ws1.append(["Category", "Requirement / Rule", "Description"])
    ws1.append(["Tester Role", "Manual Tester", "End-to-end verification"])
    ws1.append(["Bug Reporting", "Standard Template", "Feature ID, Scenario, ..."])

    ws2 = wb.create_sheet("Feature Rules")
    ws2.append(["Feature", "Logic Item", "Rule / Condition", "Expected Behavior"])
    ws2.append(
        ["Heart System", "Passive Regen", "Count < Max (5)", "1 Heart every 20 mins"]
    )
    ws2.append(
        [
            "Heart System",
            "Consumption",
            "Fail / Exit / Restart",
            "Deduct 1 Heart after popup",
        ]
    )
    ws2.append(["Royal Pass", "EXP Progression", "Win 1 Level", "Grant +1 EXP exactly"])
    ws2.append(["Lava Quest", "Streak Logic", "7 Level Wins", "Must be consecutive"])

    ws3 = wb.create_sheet("Edge Cases")
    ws3.append(["ID", "Scenario", "Condition", "Required Handling (Recovery)"])
    ws3.append(
        [
            "HEART-EDG-01",
            "Clock Manipulation",
            "Set clock back",
            "Server time must override",
        ]
    )
    ws3.append(["HEART-EDG-02", "App Kill Mid-level", "Force close", "Deduct 1 Heart"])

    ws4 = wb.create_sheet("Release Checklist")
    ws4.append(["Check ID", "Verify Item", "Pass/Fail", "Notes"])
    ws4.append(["CHK-01", "Hearts not deducted on Win", "", ""])
    ws4.append(["CHK-02", "Gold Pass updates Heart Max to 8", "", ""])
    ws4.append(["CHK-07", "Cooldown screen visible for 10 mins", "", ""])

    path = tmp_path / "test_rules.xlsx"
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# Parser tests                                                                #
# --------------------------------------------------------------------------- #


class TestReadRulesExcel:
    def test_parses_all_four_sheets(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        assert len(doc.guidelines) == 2
        assert len(doc.feature_rules) == 4
        assert len(doc.edge_cases) == 2
        assert len(doc.checklist) == 3

    def test_parses_feature_rule_fields(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        rule = doc.feature_rules[0]
        assert rule.feature == "Heart System"
        assert rule.logic_item == "Passive Regen"
        assert rule.condition == "Count < Max (5)"
        assert rule.expected == "1 Heart every 20 mins"

    def test_parses_edge_case_fields(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        edge = doc.edge_cases[0]
        assert edge.edge_id == "HEART-EDG-01"
        assert edge.scenario == "Clock Manipulation"
        assert edge.condition == "Set clock back"

    def test_parses_checklist_fields(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        check = doc.checklist[0]
        assert check.check_id == "CHK-01"
        assert check.description == "Hearts not deducted on Win"

    def test_handles_file_not_found(self):
        with pytest.raises(FileNotFoundError):
            read_rules_excel("/nonexistent/path.xlsx")


# --------------------------------------------------------------------------- #
# Resolve suites tests                                                        #
# --------------------------------------------------------------------------- #


class TestResolveSuites:
    def test_all_mapped_rules_produce_suites(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        suites = resolve_suites(doc)
        ids = [s.suite_id for s in suites]
        assert "HEART_PassiveRegen" in ids
        assert "HEART_Consumption" in ids
        assert "ROYALPASS_EXP" in ids
        assert "LAVAQUEST_Streak" in ids

    def test_unmapped_rule_raises_error(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        doc.feature_rules.append(
            FeatureRule(
                feature="UnknownFeature",
                logic_item="UnknownRule",
                condition="?",
                expected="?",
            )
        )
        with pytest.raises(
            AirtestError,
            match="Unmapped rule logic item: 'UnknownRule' in feature 'UnknownFeature'",
        ):
            resolve_suites(doc)

    def test_unmapped_edge_raises_error(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        doc.edge_cases.append(
            EdgeCase(
                edge_id="UNKNOWN-EDG-99",
                scenario="Unknown edge",
                condition="?",
                recovery="?",
            )
        )
        with pytest.raises(
            AirtestError, match=r"Unmapped edge case: 'UNKNOWN-EDG-99' \(Unknown edge\)"
        ):
            resolve_suites(doc)

    def test_unmapped_checklist_raises_error(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        doc.checklist.append(
            ChecklistItem(check_id="CHK-99", description="Unknown check")
        )
        with pytest.raises(
            AirtestError, match=r"Unmapped checklist item: 'CHK-99' \(Unknown check\)"
        ):
            resolve_suites(doc)

    def test_edge_cases_produce_suites(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        suites = resolve_suites(doc)
        ids = [s.suite_id for s in suites]
        assert "EDGE_HEART_ClockManipulation" in ids
        assert "EDGE_HEART_AppKill" in ids

    def test_checklist_produces_suites(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        suites = resolve_suites(doc)
        ids = [s.suite_id for s in suites]
        assert "CHECK_HeartNoDeductOnWin" in ids
        assert "CHECK_GoldPassMax8" in ids
        assert "CHECK_CooldownScreen" in ids

    def test_suite_steps_have_correct_suite_id(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        suites = resolve_suites(doc)
        for sd in suites:
            for step in sd.steps:
                assert step.suite_id == sd.suite_id

    def test_suite_steps_have_incremental_step_numbers(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        suites = resolve_suites(doc)
        for sd in suites:
            for i, step in enumerate(sd.steps, start=1):
                assert step.step_no == i

    def test_consumption_suite_has_specific_steps(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        suites = resolve_suites(doc)
        consumption = next(s for s in suites if s.suite_id == "HEART_Consumption")
        actions = [s.action for s in consumption.steps]
        assert actions == ["START_APP", "TAP", "WAIT_FOR", "TAP", "ASSERT_VISIBLE"]

    def test_resolve_empty_doc_returns_no_suites(self):
        doc = RulesDoc()
        suites = resolve_suites(doc)
        assert suites == []

    def test_global_assets_resolved(self, rules_wb_file):
        doc = read_rules_excel(str(rules_wb_file))
        suites = resolve_suites(doc)
        from rules_to_suites import _resolve_assets

        assets = _resolve_assets(suites)
        assert isinstance(assets, dict)
        assert len(assets) > 0


# --------------------------------------------------------------------------- #
# Generation tests                                                            #
# --------------------------------------------------------------------------- #


class TestGenerateAll:
    def test_generates_air_scripts_that_compile(self, rules_wb_file, tmp_path):
        doc = read_rules_excel(str(rules_wb_file))
        suites = resolve_suites(doc)
        gen = AirtestGenerator()
        written, issues = generate_all(
            suites,
            gen,
            output_dir=str(tmp_path),
            plan="rules_test",
            app_package="com.test",
        )
        assert len(written) == len(suites)
        for sid, path in written:
            assert os.path.isfile(path)
            py_compile.compile(path, doraise=True)

    def test_each_suite_gets_separate_air_dir(self, rules_wb_file, tmp_path):
        doc = read_rules_excel(str(rules_wb_file))
        suites = resolve_suites(doc)
        gen = AirtestGenerator()
        written, _ = generate_all(
            suites,
            gen,
            output_dir=str(tmp_path),
            plan="rules_test",
            app_package="com.test",
        )
        dirnames = set(os.path.basename(os.path.dirname(p)) for _, p in written)
        assert len(dirnames) == len(suites)

    def test_generated_script_contains_suite_id(self, rules_wb_file, tmp_path):
        doc = read_rules_excel(str(rules_wb_file))
        suites = resolve_suites(doc)
        gen = AirtestGenerator()
        written, _ = generate_all(
            suites,
            gen,
            output_dir=str(tmp_path),
            plan="rules_test",
            app_package="com.test",
        )
        for sid, path in written:
            content = open(path, encoding="utf-8").read()
            assert sid in content

    def test_generates_with_empty_app_package(self, rules_wb_file, tmp_path):
        doc = read_rules_excel(str(rules_wb_file))
        suites = resolve_suites(doc)
        gen = AirtestGenerator()
        written, issues = generate_all(
            suites,
            gen,
            output_dir=str(tmp_path),
            plan="rules",
            app_package="",
        )
        assert len(written) == len(suites)


# --------------------------------------------------------------------------- #
# inject_config tests                                                         #
# --------------------------------------------------------------------------- #


class TestInjectConfig:
    def test_base_inject_config_emits_comment(self):
        gen = AirtestGenerator()
        lines = gen.inject_config({"heart": 4, "unlimited": True})
        assert len(lines) == 1
        assert lines[0].startswith("# config:")

    def test_start_app_calls_inject_config(self):
        gen = AirtestGenerator()
        step = Step(
            suite_id="S",
            step_no=1,
            action="START_APP",
            params='{"heart":4}',
            excel_row=2,
        )
        ctx = GenCtx(assets={}, app_package="com.test")
        lines, issue = _HANDLERS["START_APP"](gen, step, ctx)
        assert issue is None
        assert any("# config" in l for l in lines)

    def test_inject_config_comment_contains_config_data(self):
        gen = AirtestGenerator()
        cfg = {"heart": 4, "unlimited": True}
        lines = gen.inject_config(cfg)
        assert "heart" in lines[0]
        assert "4" in lines[0]

    def test_subclass_inject_config_emits_custom_lines(self):
        class CustomConfigGen(AirtestGenerator):
            def inject_config(self, cfg: dict) -> list[str]:
                return [f"my_cheat.set_{k}({v!r})" for k, v in cfg.items()]

        gen = CustomConfigGen()
        lines = gen.inject_config({"heart": 4, "unlimited": True})
        assert "my_cheat.set_heart(4)" in lines
        assert "my_cheat.set_unlimited(True)" in lines
        assert len(lines) == 2

    def test_subclass_start_app_uses_custom_inject_config(self):
        class CustomConfigGen(AirtestGenerator):
            def inject_config(self, cfg: dict) -> list[str]:
                return [f"cheat.set_{k}({v!r})" for k, v in cfg.items()]

        gen = CustomConfigGen()
        step = Step(
            suite_id="S",
            step_no=1,
            action="START_APP",
            params='{"heart":4,"pass":"gold"}',
            excel_row=2,
        )
        ctx = GenCtx(assets={}, app_package="com.test")
        lines, issue = gen._HANDLERS["START_APP"](gen, step, ctx)
        assert issue is None, issue.reason if issue else ""
        assert "cheat.set_heart(4)" in lines
        assert "cheat.set_pass('gold')" in lines

    def test_inject_config_empty_dict_emits_comment(self):
        gen = AirtestGenerator()
        lines = gen.inject_config({})
        assert len(lines) == 1
        assert lines[0].startswith("# config:")

    def test_inject_config_numeric_value(self):
        gen = AirtestGenerator()
        lines = gen.inject_config({"streak": 0})
        assert any("0" in l for l in lines)

    def test_start_app_invalid_params_still_fails(self):
        gen = AirtestGenerator()
        step = Step(
            suite_id="S", step_no=1, action="START_APP", params="not-json", excel_row=2
        )
        ctx = GenCtx(assets={}, app_package="com.test")
        lines, issue = _HANDLERS["START_APP"](gen, step, ctx)
        assert issue is not None
        assert "INVALID_PARAMS_JSON" in issue.reason


# --------------------------------------------------------------------------- #
# Diagnostic hints for new error codes                                        #
# --------------------------------------------------------------------------- #


class TestDiagnosticHints:
    def test_server_validation_hint(self):
        from excel_to_airtest import _diagnostic_hint

        hint = _diagnostic_hint("SERVER_VALIDATION")
        assert "server" in hint.lower()
        assert "manual" in hint.lower()

    def test_manual_timer_check_hint(self):
        from excel_to_airtest import _diagnostic_hint

        hint = _diagnostic_hint("MANUAL_TIMER_CHECK")
        assert "timer" in hint.lower()
        assert "manual" in hint.lower()


# --------------------------------------------------------------------------- #
# CLI smoke test                                                              #
# --------------------------------------------------------------------------- #


class TestCLI:
    def test_cli_help_returns_success(self):
        import subprocess

        SCRIPT = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "rules_to_suites.py",
        )
        result = subprocess.run(
            [sys.executable, SCRIPT, "--help"],
            capture_output=True,
            text=True,
        )
        assert result.returncode == 0
        assert "excel_file" in result.stdout

    def test_cli_generates_air_scripts(self, rules_wb_file, tmp_path):
        import subprocess

        SCRIPT = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "rules_to_suites.py",
        )
        out = tmp_path / "out"
        result = subprocess.run(
            [
                sys.executable,
                SCRIPT,
                str(rules_wb_file),
                "--output",
                str(out),
                "--plan",
                "r",
                "--app-package",
                "com.demo",
            ],
            capture_output=True,
            text=True,
        )
        assert result.returncode == 0, result.stdout + result.stderr
        assert os.path.isdir(out / "r")
        py_files = list((out / "r").rglob("*.py"))
        assert len(py_files) >= 4

    def test_cli_report_flag(self, rules_wb_file, tmp_path):
        import subprocess

        SCRIPT = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "rules_to_suites.py",
        )
        out = tmp_path / "out"
        result = subprocess.run(
            [
                sys.executable,
                SCRIPT,
                str(rules_wb_file),
                "--output",
                str(out),
                "--plan",
                "r",
                "--app-package",
                "com.demo",
                "--report",
            ],
            capture_output=True,
            text=True,
        )
        assert result.returncode == 0
        report = out / "generation_report_rules.txt"
        assert report.is_file()

    def test_cli_list_projects(self, rules_wb_file, tmp_path):
        import subprocess

        SCRIPT = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "rules_to_suites.py",
        )
        result = subprocess.run(
            [sys.executable, SCRIPT, "--list-projects"],
            capture_output=True,
            text=True,
        )
        assert result.returncode == 0
        assert "pixon" in result.stdout
