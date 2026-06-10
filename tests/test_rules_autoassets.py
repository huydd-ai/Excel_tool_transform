"""Regression tests for rules_to_suites asset auto-derivation (Part B refactor).

Per-suite assets are derived from step targets via _assets_from_targets, not
declared manually in each builder.
"""

import openpyxl
import pytest

from rules_to_suites import read_rules_excel, resolve_suites, _resolve_assets


@pytest.fixture
def rules_wb_file(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Feature Rules"
    ws1.append(["Feature", "Logic Item", "Rule / Condition", "Expected Behavior"])
    ws1.append(["Heart System", "Consumption", "Fail / Exit", "Deduct 1 Heart"])
    ws1.append(["Heart System", "Passive Regen", "Count < Max", "1 Heart / 20 min"])
    wb.save(tmp_path / "r.xlsx")
    return tmp_path / "r.xlsx"


def _suite(suites, sid):
    return next(s for s in suites if s.suite_id == sid)


def test_consumption_assets_derived_from_targets(rules_wb_file):
    suites = resolve_suites(read_rules_excel(str(rules_wb_file)))
    consumption = _suite(suites, "HEART_Consumption")
    assert set(consumption.assets) == {
        "btn_play",
        "confirm_popup",
        "btn_confirm",
        "heart_4",
    }


def test_suite_without_targets_has_no_assets(rules_wb_file):
    # Passive Regen has only START_APP + _todo_step (no image targets).
    suites = resolve_suites(read_rules_excel(str(rules_wb_file)))
    passive = _suite(suites, "HEART_PassiveRegen")
    assert passive.assets == {}


def test_derived_assets_come_from_base_catalog(rules_wb_file):
    from rules_to_suites import _BASE_ASSETS

    suites = resolve_suites(read_rules_excel(str(rules_wb_file)))
    consumption = _suite(suites, "HEART_Consumption")
    for name, asset in consumption.assets.items():
        assert asset is _BASE_ASSETS[name]


def test_combined_assets_nonempty(rules_wb_file):
    suites = resolve_suites(read_rules_excel(str(rules_wb_file)))
    combined = _resolve_assets(suites)
    assert "btn_play" in combined
    assert len(combined) > 0


def test_positional_fallback_when_headers_renamed(tmp_path):
    # Headers don't match any known name -> _col falls back to fixed positions.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feature Rules"
    ws.append(["ColA", "ColB", "ColC", "ColD"])  # unrecognized headers
    ws.append(["Heart System", "Consumption", "Fail / Exit", "Deduct 1 Heart"])
    path = tmp_path / "renamed.xlsx"
    wb.save(path)
    doc = read_rules_excel(str(path))
    rule = doc.feature_rules[0]
    assert rule.feature == "Heart System"
    assert rule.logic_item == "Consumption"
    assert rule.condition == "Fail / Exit"
    assert rule.expected == "Deduct 1 Heart"


def test_header_mapping_survives_column_reorder(tmp_path):
    # Columns shuffled but headers present -> header-map must still parse correctly.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feature Rules"
    ws.append(["Expected Behavior", "Logic Item", "Feature", "Rule / Condition"])
    ws.append(["Deduct 1 Heart", "Consumption", "Heart System", "Fail / Exit"])
    path = tmp_path / "reordered.xlsx"
    wb.save(path)
    doc = read_rules_excel(str(path))
    rule = doc.feature_rules[0]
    assert rule.feature == "Heart System"
    assert rule.logic_item == "Consumption"
    assert rule.expected == "Deduct 1 Heart"
