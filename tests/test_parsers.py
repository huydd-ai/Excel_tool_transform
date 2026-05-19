"""Parsers: Object_Repository / Test_Execution / Action_Logic."""
import openpyxl

from excel_to_airtest import (
    LOCATOR_IMAGE,
    LOCATOR_OCR,
    parse_action_logic,
    parse_object_repository,
    parse_test_execution,
)


# --------------------------------------------------------------------------- #
# Object_Repository                                                           #
# --------------------------------------------------------------------------- #

def test_object_repo_loads_assets_with_typed_fields(make_wb_file):
    p = make_wb_file(objects=[
        ["btn_play",  "IMAGE", "./assets/btn.png", 0.85, 3],
        ["heart_txt", "OCR",   "NONE",             0.7,  10],
    ])
    wb = openpyxl.load_workbook(p, data_only=True)
    assets, errs = parse_object_repository(wb, "Object_Repository")

    assert errs == []
    assert assets["btn_play"].locator_type == LOCATOR_IMAGE
    assert assets["btn_play"].resource_path == "./assets/btn.png"
    assert assets["btn_play"].threshold == 0.85
    assert assets["btn_play"].timeout == 3
    assert assets["heart_txt"].locator_type == LOCATOR_OCR


def test_object_repo_applies_defaults_for_missing_fields(make_wb_file):
    p = make_wb_file(objects=[["only_id", None, None, None, None]])
    wb = openpyxl.load_workbook(p, data_only=True)
    assets, _ = parse_object_repository(wb, "Object_Repository")

    a = assets["only_id"]
    assert a.locator_type == LOCATOR_IMAGE
    assert a.resource_path == ""
    assert a.threshold == 0.8
    assert a.timeout == 5.0


def test_object_repo_skips_blank_rows_and_empty_ids(make_wb_file):
    p = make_wb_file(objects=[
        ["a",  "IMAGE", "x.png", 0.8, 1],
        [None, None,    None,    None, None],
        ["",   "IMAGE", "y.png", 0.8, 1],
        ["b",  "IMAGE", "z.png", 0.8, 1],
    ])
    wb = openpyxl.load_workbook(p, data_only=True)
    assets, _ = parse_object_repository(wb, "Object_Repository")
    assert set(assets) == {"a", "b"}


def test_object_repo_returns_error_when_sheet_missing():
    wb = openpyxl.Workbook()
    assets, errs = parse_object_repository(wb, "Nope")
    assert assets == {}
    assert errs and "not found" in errs[0]


def test_object_repo_returns_error_when_required_column_missing(make_wb_file):
    p = make_wb_file(
        objects=[["x", "y"]],
        obj_headers=["Foo", "Bar"],
    )
    wb = openpyxl.load_workbook(p, data_only=True)
    _, errs = parse_object_repository(wb, "Object_Repository")
    assert errs and "object_id" in errs[0]


def test_object_repo_headers_are_case_insensitive(make_wb_file):
    p = make_wb_file(
        objects=[["x", "IMAGE", "a.png", 0.9, 7]],
        obj_headers=["object_ID", "LOCATOR_type", "Resource_Path", "smart_THRESHOLD", "TIMEOUT"],
    )
    wb = openpyxl.load_workbook(p, data_only=True)
    assets, errs = parse_object_repository(wb, "Object_Repository")

    assert errs == []
    assert assets["x"].threshold == 0.9
    assert assets["x"].timeout == 7


# --------------------------------------------------------------------------- #
# Test_Execution                                                              #
# --------------------------------------------------------------------------- #

def test_test_execution_groups_by_suite(make_wb_file):
    p = make_wb_file(steps=[
        ["S1", 1, "CLICK",    "a", "", "ok"],
        ["S1", 2, "WAIT_FOR", "b", "", "ok"],
        ["S2", 1, "CLICK",    "c", "", "ok"],
    ])
    wb = openpyxl.load_workbook(p, data_only=True)
    suites, errs = parse_test_execution(wb, "Test_Execution")

    assert errs == []
    assert set(suites) == {"S1", "S2"}
    assert len(suites["S1"]) == 2
    assert len(suites["S2"]) == 1


def test_test_execution_tracks_excel_row_starting_at_2(make_wb_file):
    p = make_wb_file(steps=[
        ["S1", 1, "CLICK",    "a", "", ""],
        ["S1", 2, "WAIT_FOR", "b", "", ""],
        ["S2", 1, "CLICK",    "c", "", ""],
    ])
    wb = openpyxl.load_workbook(p, data_only=True)
    suites, _ = parse_test_execution(wb, "Test_Execution")

    # Row 1 is header; first data row is 2.
    assert suites["S1"][0].excel_row == 2
    assert suites["S1"][1].excel_row == 3
    assert suites["S2"][0].excel_row == 4


def test_test_execution_uppercases_action_keyword(make_wb_file):
    p = make_wb_file(steps=[["S", 1, "click", "a", "", ""]])
    wb = openpyxl.load_workbook(p, data_only=True)
    suites, _ = parse_test_execution(wb, "Test_Execution")
    assert suites["S"][0].action == "CLICK"


def test_test_execution_skips_blank_and_suiteless_rows(make_wb_file):
    p = make_wb_file(steps=[
        ["S1", 1, "CLICK", "a", "", ""],
        [None, None, None, None, None, None],
        ["",   2, "CLICK", "b", "", ""],
        ["S1", 3, "CLICK", "c", "", ""],
    ])
    wb = openpyxl.load_workbook(p, data_only=True)
    suites, _ = parse_test_execution(wb, "Test_Execution")

    assert set(suites) == {"S1"}
    assert [s.target for s in suites["S1"]] == ["a", "c"]


def test_test_execution_returns_error_when_required_column_missing(make_wb_file):
    p = make_wb_file(
        steps=[["S1", 1]],
        step_headers=["Suite_ID", "Step"],
    )
    wb = openpyxl.load_workbook(p, data_only=True)
    _, errs = parse_test_execution(wb, "Test_Execution")
    assert errs and "action_keyword" in errs[0]


# --------------------------------------------------------------------------- #
# Action_Logic                                                                #
# --------------------------------------------------------------------------- #

def test_action_logic_splits_act_primitives_and_flow_docs(make_wb_file):
    p = make_wb_file(actions=[
        ["ACT_001",  "CLICK",          "CLICK(Object_ID)",    ""],
        ["ACT_002",  "wait_for",       "WAIT_FOR(Object_ID)", ""],
        ["FLOW_001", "Display Splash", "narrative",           "SplashScreen"],
    ])
    wb = openpyxl.load_workbook(p, data_only=True)
    kws, flows, errs = parse_action_logic(wb, "Action_Logic")

    assert errs == []
    assert kws == {"CLICK", "WAIT_FOR"}
    assert len(flows) == 1
    assert flows[0].logic_id == "FLOW_001"
    assert flows[0].target_page == "SplashScreen"


def test_action_logic_silent_when_optional_sheet_missing():
    wb = openpyxl.Workbook()
    kws, flows, errs = parse_action_logic(wb, "Action_Logic")
    assert kws == set()
    assert flows == []
    assert errs == []
