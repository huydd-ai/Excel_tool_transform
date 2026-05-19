"""--init-excel generates a valid template.xlsx with 3 sheets and action dropdown."""
import os
import sys

import openpyxl
import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from templates.excel_template import ALL_KEYWORDS, generate_template


def test_generate_template_creates_file(tmp_path):
    out = generate_template(str(tmp_path / "template.xlsx"))
    assert os.path.isfile(out)


def test_template_has_three_sheets(tmp_path):
    out = generate_template(str(tmp_path / "template.xlsx"))
    wb = openpyxl.load_workbook(out)
    assert set(wb.sheetnames) == {"Object_Repository", "Action_Logic", "Test_Execution"}


def test_object_repository_headers(tmp_path):
    out = generate_template(str(tmp_path / "template.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb["Object_Repository"]
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == ["Object_ID", "Locator_Type", "Resource_Path", "Smart_Threshold", "Timeout"]


def test_object_repository_has_example_rows(tmp_path):
    out = generate_template(str(tmp_path / "template.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb["Object_Repository"]
    ids = [ws.cell(r, 1).value for r in range(2, 10) if ws.cell(r, 1).value]
    assert len(ids) >= 2


def test_test_execution_headers(tmp_path):
    out = generate_template(str(tmp_path / "template.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb["Test_Execution"]
    headers = [ws.cell(1, c).value for c in range(1, 7)]
    assert headers == ["Suite_ID", "Step", "Action_Keyword", "Target_ID", "Params", "Expected_Result"]


def test_test_execution_has_example_rows(tmp_path):
    out = generate_template(str(tmp_path / "template.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb["Test_Execution"]
    suite_vals = [ws.cell(r, 1).value for r in range(2, 10) if ws.cell(r, 1).value]
    assert len(suite_vals) >= 3


def test_action_logic_has_all_keywords(tmp_path):
    out = generate_template(str(tmp_path / "template.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb["Action_Logic"]
    action_names = [ws.cell(r, 2).value for r in range(2, 30) if ws.cell(r, 2).value]
    for kw in ALL_KEYWORDS:
        assert kw in action_names, f"Keyword '{kw}' missing from Action_Logic"


def test_test_execution_has_data_validation(tmp_path):
    out = generate_template(str(tmp_path / "template.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb["Test_Execution"]
    validations = list(ws.data_validations.dataValidation)
    assert len(validations) >= 1
    dv = validations[0]
    assert dv.type == "list"
    for kw in ["TAP", "TOUCH", "SLEEP"]:
        assert kw in (dv.formula1 or "")
